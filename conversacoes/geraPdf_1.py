import os, openpyxl,  shutil
from datetime import datetime
from copy import copy
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import win32com.client as win32


from conversacoes.FunComplementar import (
    formatar_cnpj_cpf_cep
)

bot_username = "PDFGen_bot"


def caminho_salvar_pdf(user_id, nome_pdf): # caminho_salvar_pdf(user_id, "orçamento - nº") e retorna o caminho da dele
    user_id = str(user_id)
    caminho_atual = os.path.abspath(__file__)
    diretorio_conversacoes = os.path.dirname(caminho_atual)
    diretorio_user_data = os.path.join(diretorio_conversacoes, "..", "user_data")
    diretorio_ususario = os.path.join(diretorio_user_data, user_id)
    caminho_arquivo = os.path.join(diretorio_ususario, "orçamentos")
    diretorio_orcamentos = os.path.join(caminho_arquivo, nome_pdf)
    return diretorio_orcamentos

def caminho_foto_usuario(user_id): # caminho_foto_usuario(user_id) e retorna o caminho da foto
    user_id = str(user_id)
    caminho_atual = os.path.abspath(__file__)
    diretorio_conversacoes = os.path.dirname(caminho_atual)
    diretorio_user_data = os.path.join(diretorio_conversacoes, "..", "user_data")
    diretorio_ususario = os.path.join(diretorio_user_data, user_id)
    caminho_arquivo = os.path.join(diretorio_ususario, user_id + ".jpg")
    return caminho_arquivo

def caminho_layout(nome_layout): # caminho_layout("nome do layout") e retorna o caminho dele
    caminho_atual = os.path.abspath(__file__)
    diretorio_conversacoes = os.path.dirname(caminho_atual)
    diretorio_user_data = os.path.join(diretorio_conversacoes, "..", "layout_excel")
    diretorio_layout = os.path.join(diretorio_user_data, nome_layout)
    return diretorio_layout




# Função para copiar o estilo de uma célula para outra
def copy_cell_style(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.border = copy(src_cell.border)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.alignment = copy(src_cell.alignment)

# Função para criar a estrutura da linha para o texto
def criar_estrutura_linha_texto(sheet, linha):
    for col in range(2, sheet.max_column + 1):
        cell = sheet.cell(row=linha, column=col)
        cell.font = Font(name="Consolas", size=10, color="000000", bold=False)
        cell.border = Border(left=Side(style="thin", color="0000FF"),
                             right=Side(style="thin", color="0000FF"),
                             top=Side(style="thin", color="0000FF"),
                             bottom=Side(style="thin", color="0000FF"))
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    # Definir a altura da linha para 53 pixels
    sheet.row_dimensions[linha].height = 43

def run_vba_in_excel(file_path, vba_code, output_pdf):
    print("path ", file_path)
    print("run ", output_pdf)
    # Iniciar o Excel
    excel = win32.Dispatch("Excel.Application")
    try:
        # Abrir o arquivo Excel
        wb = excel.Workbooks.Open(file_path)
        # Acessar o módulo VBA do arquivo
        vba_module = wb.VBProject.VBComponents.Add(1)  # 1 significa adicionar um módulo normal
        # Inserir o código VBA no módulo
        vba_module.CodeModule.AddFromString(vba_code)
        # Executar o código VBA
        excel.Application.Run("Macro1", output_pdf)  # Chamar a macro "Macro1" e passar o nome do arquivo PD
        # Salvar e fechar o arquivo
        wb.Save()
        wb.Close()
    except Exception as e:
        print("Erro:", e)
    finally:
        # Fechar o Excel
        excel.Quit()

def preencher_dados_excel(json_data, current_directory):
    # caminho_salvar_pdf(user_id, "orçamento - nº") e retorna o caminho da dele
    # caminho_foto_usuario(user_id) e retorna o caminho da foto
    # caminho_layout("nome do layout") e retorna o caminho dele
    excel_path = caminho_layout("planilha.xlsx")
    temp_excel_path = os.path.join(os.path.dirname(excel_path), "temp_" + os.path.basename(excel_path))
    planilha_path = temp_excel_path
    
    shutil.copyfile(excel_path, planilha_path)
    wb = openpyxl.load_workbook(planilha_path)
    sheet = wb["Layout_orcamento"]

    # Preencher dados da matriz cadastrada
    matriz = json_data.get("matriz", {})
    sheet["B1"] = matriz.get("nome", "")
    sheet["B2"] = formatar_cnpj_cpf_cep(matriz.get("cnpj_cpf", ""))
    sheet["B3"] = matriz.get("endereco", "")
    sheet["B4"] = f"{matriz.get('bairro', '')}, {matriz.get('cidade', '')} - {matriz.get('estado', '')}, {matriz.get('cep', '')}"
    sheet["B5"] = f"E-mail: {matriz.get('email', '')} - Tel.: {matriz.get('telefone_contato', '')}"
    sheet["F1"] = f"Orçamento - {json_data.get('numero_orcamento', '')}"
    sheet["G5"] = json_data.get("data_emissao", "")
    sheet["G6"] = json_data.get("data_vencimento", "")

    # Preencher dados do cliente selecionado
    cliente = json_data.get("cliente", {})
    sheet["B7"] = cliente.get("nome", "")
    sheet["B8"] = formatar_cnpj_cpf_cep(cliente.get("cnpj_cpf", ""))
    sheet["B9"] = cliente.get("endereco", "")
    sheet["B10"] = f"{cliente.get('bairro', '')}, {cliente.get('cidade', '')} - {cliente.get('estado', '')}, {cliente.get('cep', '')}"
    sheet["B11"] = f"E-mail: {cliente.get('email', '')} - Tel.: {cliente.get('telefone_contato', '')}"

    # Preencher textos
    orcamento_mensagem = json_data.get("orcamento_mensagem", [])
    linha_atual = 14
    for mensagem in orcamento_mensagem:

        texto = mensagem.get("texto", "")
        # Copiar a estrutura da linha 14 para a nova linha
        sheet.insert_rows(linha_atual, amount=1)
        criar_estrutura_linha_texto(sheet, linha_atual)
        sheet.merge_cells(start_row=linha_atual, start_column=2, end_row=linha_atual, end_column=sheet.max_column)
        sheet.cell(row=linha_atual, column=2, value=texto)
        linha_atual += 1
    # Preencher valor total
    totalk = json_data.get("valor_total", "")
    try:
        sheet[f"G{(linha_atual + 1)}"] = int(totalk)
    except ValueError:
        # Caso a conversão para inteiro falhe, tentar converter para um número de ponto flutuante (float)
        try:
            sheet[f"G{(linha_atual + 1)}"] = float(totalk)
        except ValueError:
            sheet[f"G{(linha_atual + 1)}"] = totalk
    # Preencher observação
    observacao = json_data.get("observacao", "Obrigado pela preferência!")
    linha_atual += 6
    # Iterar sobre as colunas da B até a G
    for col in range(2, sheet.max_column + 1):
        src_cell = sheet.cell(row=14, column=col)
        dst_cell = sheet.cell(row=linha_atual, column=col)
        
        # Definir o valor da célula como a observacao
        dst_cell.value = observacao
        
        # Definir a quebra de texto para True
        dst_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Definir a fonte Century Gothic e tamanho 10
        font = Font(name='Century Gothic', size=10)
        dst_cell.font = font

    # Mesclar as células da linha atual da coluna B até a coluna G
    sheet.merge_cells(start_row=linha_atual, start_column=2, end_row=linha_atual, end_column=sheet.max_column)

    # Definir a altura da linha para acomodar o texto quebrado
    sheet.row_dimensions[linha_atual].height = 153
    # Obter a data e hora atual
    
    # Formatar a data e hora no formato desejado: dd/mm/yyyy hh:mm
    data_hora_formatada = datetime.now().strftime("%d/%m/%Y às %H:%M")
    rodape = f"O orçamento emitido pelo https://t.me/{bot_username} no Telegram data de emissão {data_hora_formatada}."
    # Definir a fórmula para a data e hora formatada
    formula_data_hora_formatada = rodape
    # Inserir a fórmula no rodapé da planilha e alinhar à esquerda
    sheet.oddFooter.right.text = formula_data_hora_formatada
    # Salvar a planilha temporariamente em um arquivo temporário
    temp_file = "temps.xlsx"
    wb.save(planilha_path)
    

    vba_code = '''
    ' Macro1 Macro
    '
    '
    Sub Macro1(outputPdf As String)
        Dim pdfFile As String
        Dim ws As Worksheet

        ' Obter o caminho da pasta do arquivo Excel e o nome do arquivo PDF a partir do argumento recebido do Python
        pdfFile = outputPdf

        ' Imprima a planilha ativa para o arquivo PDF
        Set ws = ActiveSheet
        ws.ExportAsFixedFormat Type:=xlTypePDF, Filename:=pdfFile, Quality:=xlQualityStandard, IncludeDocProperties:=True, IgnorePrintAreas:=False

    End Sub
    ''' 
        # Obter o diretório atual do script Python


    run_vba_in_excel(planilha_path, vba_code, current_directory)

    os.remove(planilha_path)