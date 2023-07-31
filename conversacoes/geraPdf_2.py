import os, shutil, openpyxl
import win32com.client as win32
from openpyxl.drawing.image import Image
from datetime import datetime
from copy import copy
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles import NamedStyle



from conversacoes.FunComplementar import (
    formatar_cnpj_cpf_cep
)

bot_username = "PDFGen_bot"



#    layout tem que ter o nome -> planilha_foto.xlsx na pasta layout_excel/
#    user_id = "381043536"
#    nome_pdf = "orçamento - 02.pdf"
#    data = "dados do .json"
#    inserir_imagem_excel(user_id, nome_pdf, data)
#    retorna o caminho do PDF
    

def criar_estrutura_linha_texto(sheet, linha):
    for col in range(3, 9):
        cell = sheet.cell(row=linha, column=col)
        cell.font = Font(name="Consolas", size=10, color="000000", bold=False)
        cell.border = Border(left=Side(style="thin", color="0000FF"),
                             right=Side(style="thin", color="0000FF"),
                             top=Side(style="thin", color="0000FF"),
                             bottom=Side(style="thin", color="0000FF"))
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    # Definir a altura da linha para 53 pixels
    sheet.row_dimensions[linha].height = 43


def caminho_salvar_pdf(user_id, nome_pdf): # caminho_salvar_pdf(user_id, "orçamento - nº") e retorna o caminho da dele
    user_id = user_id
    caminho_atual = os.path.abspath(__file__)
    diretorio_conversacoes = os.path.dirname(caminho_atual)
    diretorio_user_data = os.path.join(diretorio_conversacoes, "..", "user_data")
    diretorio_ususario = os.path.join(diretorio_user_data, user_id)
    caminho_arquivo = os.path.join(diretorio_ususario, "orçamentos")
    diretorio_orcamentos = os.path.join(caminho_arquivo, nome_pdf)
    return diretorio_orcamentos

def caminho_foto_usuario(user_id): # caminho_foto_usuario(user_id) e retorna o caminho da foto
    user_id = user_id
    caminho_atual = os.path.abspath(__file__)
    diretorio_conversacoes = os.path.dirname(caminho_atual)
    diretorio_user_data = os.path.join(diretorio_conversacoes, "..", "user_data")
    diretorio_ususario = os.path.join(diretorio_user_data, user_id)
    caminho_arquivo = os.path.join(diretorio_ususario, user_id + ".jpg")
    return caminho_arquivo

def caminho_layout(nome_layout): # caminho_layout("nome do layout") e retorna o caminho dele
    caminho_atual = os.path.abspath(__file__)
    diretorio_conversacoes = os.path.dirname(caminho_atual)
    diretorio_user_data = os.path.join(diretorio_conversacoes, "..", "layout_excel")
    diretorio_layout = os.path.join(diretorio_user_data, nome_layout)
    return diretorio_layout

def inserir_imagem_excel(user_id, nome_pdf, json_data):
    
    imagem_path = caminho_foto_usuario(user_id)
    
    largura_cm = 6.44
    altura_cm = 4.15
    excel_path = caminho_layout("planilha_foto.xlsx")
    
    if not os.path.exists(imagem_path) or not os.path.exists(excel_path):
        print("Arquivo de imagem ou Excel não encontrado.")
        return

    # Converter as dimensões de centímetros para pontos (1 cm = 28.3465 pontos no Excel)
    largura_pontos = int(largura_cm * 28.3465)
    altura_pontos = int(altura_cm * 28.3465)

    # Criar uma instância do Excel
    excel = win32.gencache.EnsureDispatch('Excel.Application')

    try:
        # Copiar o arquivo Excel para um arquivo temporário
        temp_excel_path = os.path.join(os.path.dirname(excel_path), "temp_" + os.path.basename(excel_path))
        
        shutil.copyfile(excel_path, temp_excel_path)

        # Abrir o arquivo Excel
        wb = excel.Workbooks.Open(temp_excel_path)
        ws = wb.Sheets("layout_orcamento")

        # Inserir a imagem no Retângulo 1
        img = Image(imagem_path)
        proporcao = img.height / img.width

        # Verificar se a proporção da imagem é maior ou menor que a proporção do retângulo
        if proporcao >= altura_cm / largura_cm:
            # Ajustar a altura para a altura do retângulo e calcular a largura proporcionalmente
            img.height = altura_pontos
            img.width = int(altura_pontos / proporcao)
        else:
            # Ajustar a largura para a largura do retângulo e calcular a altura proporcionalmente
            img.width = largura_pontos
            img.height = int(largura_pontos * proporcao)

        # Definir a posição da imagem no retângulo
        img_left = ws.Shapes("Retângulo 1").Left
        img_top = ws.Shapes("Retângulo 1").Top
        img.left = img_left
        img.top = img_top

        # Salvar as alterações no arquivo Excel temporário
        wb.Save()

        # Fechar o arquivo Excel temporário
        wb.Close()

        caminho_salvar = caminho_salvar_pdf(user_id, nome_pdf)
        
        
        
        run_vba_in_excel_foto(temp_excel_path, imagem_path)
        
        # Excluir o arquivo Excel temporário
        
        preencher_dados_excel(json_data, caminho_salvar, temp_excel_path)
        
        
        os.remove(temp_excel_path)

        return caminho_salvar
    except Exception as e:
        print(f"Ocorreu um erro ao inserir a imagem e executar o macro: {str(e)}")

    finally:
        # Fechar o aplicativo Excel
        excel.Quit()

def run_vba_in_excel_foto(file_path, imagem_path):
    
    vba_code = '''
  
    Sub Macro2(caminho As String)
        Dim ws As Worksheet
        Set ws = ThisWorkbook.Sheets("layout_orcamento")
        ' Caminho da imagem a ser inserida
        Dim imagePath As String
        imagePath = caminho
        
        ' Definir a posição e o tamanho do controle ActiveX "Image1"
        Dim leftPosition As Double
        Dim topPosition As Double
        Dim width As Double
        Dim height As Double
        
        leftPosition = ws.Shapes("Image1").Left
        topPosition = ws.Shapes("Image1").Top
        width = ws.Shapes("Image1").Width
        height = ws.Shapes("Image1").Height
        
        ' Inserir a imagem no controle ActiveX "Image1"
        ws.Shapes.AddPicture Filename:=imagePath, linktofile:=msoFalse, _
            savewithdocument:=msoCTrue, Left:=leftPosition, Top:=topPosition, Width:=width, Height:=height

    End Sub

    ''' 
    # Iniciar o Excel
    excel = win32.Dispatch("Excel.Application")
    try:
        # Abrir o arquivo Excel
        wb = excel.Workbooks.Open(file_path)
        # Acessar o módulo VBA do arquivo
        vba_module = wb.VBProject.VBComponents.Add(1)  # 1 significa adicionar um módulo normal
        # Inserir o código VBA no módulo
        vba_module.CodeModule.AddFromString(vba_code)
        # Executar o código VBA
        excel.Application.Run("Macro2", imagem_path)  # Chamar a macro "Macro1" e passar o nome do arquivo PDF
        # Salvar e fechar o arquivo
        wb.Save()
        wb.Close()
    except Exception as e:
        print("Erro:", e)
    finally:
        # Fechar o Excel
        excel.Quit()

    #preencher_dados_excel(json_data, caminho_salvar, temp_excel_path)
def preencher_dados_excel(json_data, current_directory, caminho_layout):
    planilha_path = caminho_layout
    wb = openpyxl.load_workbook(planilha_path)
    sheet = wb["Layout_orcamento"]

    # Preencher dados da matriz cadastrada
    matriz = json_data.get("matriz", {})
    sheet["E2"] = matriz.get("nome", "asaa")
    sheet["E3"] = formatar_cnpj_cpf_cep(matriz.get("cnpj_cpf", ""))
    sheet["E4"] = f"{matriz.get('endereco', '')}"
    sheet["E5"] = f"{matriz.get('bairro', '')}, {matriz.get('cidade', '')} - {matriz.get('estado', '')}, {matriz.get('cep', '')}"
    sheet["E6"] = f"{matriz.get('email', '')}"
    sheet["E7"] = f"{matriz.get('telefone_contato', '')}"
    sheet["H2"] = f"Orçamento Nº {json_data.get('numero_orcamento', '')}"
    sheet["I6"] = json_data.get("data_emissao", "")
    sheet["I7"] = json_data.get("data_vencimento", "")

    # Preencher dados do cliente selecionado
    cliente = json_data.get("cliente", {})
    sheet["C8"] = cliente.get("nome", "")
    sheet["C9"] = formatar_cnpj_cpf_cep(cliente.get("cnpj_cpf", ""))
    sheet["C10"] = f"{cliente.get('endereco', '')}"
    sheet["C11"] = f"{cliente.get('bairro', '')}, {cliente.get('cidade', '')} - {cliente.get('estado', '')}, {cliente.get('cep', '')}"
    sheet["C12"] = f"{cliente.get('email', '')}"
    sheet["C13"] = f"{cliente.get('telefone_contato', '')}"


# Preencher textos
    orcamento_mensagem = json_data.get("orcamento_mensagem", [])
    linha_atual = 16
    for mensagem in orcamento_mensagem:

        texto = mensagem.get("texto", "")
        # Copiar a estrutura da linha 14 para a nova linha
        sheet.insert_rows(linha_atual, amount=1)
        criar_estrutura_linha_texto(sheet, linha_atual)
        sheet.merge_cells(start_row=linha_atual, start_column=3, end_row=linha_atual, end_column=9)
        sheet.cell(row=linha_atual, column=3, value=texto)
        linha_atual += 1
    # Preencher valor total
    totalk = json_data.get("valor_total", "")
    try:
       totalk = int(totalk)
    except ValueError:
        # Caso a conversão para inteiro falhe, tentar converter para um número de ponto flutuante (float)
        try:
            totalk = float(totalk)
        except ValueError:
            totalk = totalk


    # Definir um estilo com o formato desejado (vírgula como separador decimal)
    comma_style = NamedStyle(name="comma_style", number_format="#,##0.00")
        # Definir as propriedades do estilo para manter o layout da célula
    comma_style.font = Font(name='Consolas', size=14, bold=True)
    comma_style.alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(),   # Borda fina em azul na esquerda
        right=Side(style='thin', color='0000FF'), # Nenhuma borda na direita
        top=Side(style='thin', color='0000FF'),    # Borda fina em azul na parte superior
        bottom=Side(style='thin', color='0000FF'), # Borda fina em azul na parte inferior
    )
    comma_style.border = border
    sheet.row_dimensions[(linha_atual + 2)].height = 32
        # Aplicar o estilo à célula
    celula = sheet.cell(row=(linha_atual + 2), column=9, value=f"R$ {totalk:.2f}")
    celula.style = comma_style
    

        # Definir um estilo com o formato desejado (vírgula como separador decimal)
    comma_styleS = NamedStyle(name="comma_styleS", number_format="")
        # Definir as propriedades do estilo para manter o layout da célula
    comma_styleS.font = Font(name='Consolas', size=14, bold=True)
    comma_styleS.alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin', color='0000FF'),   # Borda fina em azul na esquerda
        right=Side(), # Nenhuma borda na direita
        top=Side(style='thin', color='0000FF'),    # Borda fina em azul na parte superior
        bottom=Side(style='thin', color='0000FF'), # Borda fina em azul na parte inferior
    )
    comma_styleS.border = border
    sheet.row_dimensions[(linha_atual + 2)].height = 32
    # Aplicar o estilo à célula
    celula = sheet.cell(row=(linha_atual + 2), column=8, value=str("SUBTOTAL"))
    celula.style = comma_styleS
    
    # Preencher observação
    observacao = json_data.get("observacao", "Obrigado pela preferência!")
    linha_atual += 7
    # Iterar sobre as colunas da B até a G
    for col in range(3, sheet.max_column + 1):
        src_cell = sheet.cell(row=14, column=col)
        dst_cell = sheet.cell(row=linha_atual, column=col)
        
        # Definir o valor da célula como a observacao
        dst_cell.value = observacao

        # Definir a quebra de texto para True
        dst_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Definir a fonte Century Gothic e tamanho 10
        font = Font(name='Century Gothic', size=10)
        dst_cell.font = font

    # Mesclar as células da linha atual da coluna B até a coluna G
    sheet.merge_cells(start_row=linha_atual, start_column=3, end_row=linha_atual, end_column=9)

    # Definir a altura da linha para acomodar o texto quebrado
    sheet.row_dimensions[linha_atual].height = 153
    # Obter a data e hora atual
    
    # Formatar a data e hora no formato desejado: dd/mm/yyyy hh:mm
    data_hora_formatada = datetime.now().strftime("%d/%m/%Y às %H:%M")
    rodape = f"O orçamento emitido pelo https://t.me/{bot_username} no Telegram data de emissão {data_hora_formatada}."
    # Definir a fórmula para a data e hora formatada
    formula_data_hora_formatada = rodape
    # Inserir a fórmula no rodapé da planilha e alinhar à esquerda
    sheet.oddFooter.right.text = formula_data_hora_formatada
       # Salvar a planilha temporariamente em um arquivo temporário
    temp_file = "temps.xlsx"
    wb.save(planilha_path)
    

    vba_code = '''
    ' Macro1 Macro
    '
    '
    Sub Macro1(outputPdf As String)
        Dim pdfFile As String
        Dim ws As Worksheet

        ' Obter o caminho da pasta do arquivo Excel e o nome do arquivo PDF a partir do argumento recebido do Python
        pdfFile = outputPdf

        ' Imprima a planilha ativa para o arquivo PDF
        Set ws = ActiveSheet
        ws.ExportAsFixedFormat Type:=xlTypePDF, Filename:=pdfFile, Quality:=xlQualityStandard, IncludeDocProperties:=True, IgnorePrintAreas:=False

    End Sub
    ''' 
        # Obter o diretório atual do script Python
    

    
    run_vba_in_excel(planilha_path, vba_code, current_directory)
   # os.remove(temp_file)


def run_vba_in_excel(file_path, vba_code, output_pdf):
    
    # Iniciar o Excel
    excel = win32.Dispatch("Excel.Application")
    try:
        # Abrir o arquivo Excel
        wb = excel.Workbooks.Open(file_path)
        # Acessar o módulo VBA do arquivo
        vba_module = wb.VBProject.VBComponents.Add(1)  # 1 significa adicionar um módulo normal
        # Inserir o código VBA no módulo
        vba_module.CodeModule.AddFromString(vba_code)
        # Executar o código VBA
        excel.Application.Run("Macro1", output_pdf)  # Chamar a macro "Macro1" e passar o nome do arquivo PD
        # Salvar e fechar o arquivo
        wb.Save()
        wb.Close()
    except Exception as e:
        print("Erro:", e)
    finally:
        # Fechar o Excel
        excel.Quit()